#!/usr/bin/env python3
"""
将客户信息.xlsx导入为Supabase SQL文件
生成: supabase-base-seed.sql
"""

import openpyxl
import os
import uuid

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', '..', '客户信息.xlsx')
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), '..', 'supabase-base-seed.sql')

def escape_sql(val):
    """转义SQL字符串中的单引号"""
    if val is None:
        return 'NULL'
    s = str(val).strip()
    if not s:
        return 'NULL'
    return "'" + s.replace("'", "''") + "'"

def main():
    print(f'读取Excel: {EXCEL_PATH}')
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # ========== 客户数据 ==========
    ws1 = wb['客户信息列表']
    customers = []
    seen_ids = set()

    for row in ws1.iter_rows(min_row=3, values_only=True):
        cid = row[0]
        if not cid or cid in seen_ids:
            continue
        seen_ids.add(cid)
        customers.append({
            'id': cid,
            'name': row[1],
            'legal_representative': row[2],
            'social_credit_code': row[3],
            'province': row[4],
            'city': row[5],
            'county': row[6],
            'business_type': row[7],
            'customer_type': row[8],
            'detail_address': row[9],
            'customer_grade': row[10],
            'credit_score': row[11],
            'actual_controller': row[12],
            'second_type': row[13],
        })

    print(f'客户数: {len(customers)}')

    # ========== 联系人数据 ==========
    ws2 = wb['联系人信息']
    contacts = []

    for row in ws2.iter_rows(min_row=3, values_only=True):
        cust_id = row[0]
        name = row[1]
        if not cust_id or not name or cust_id not in seen_ids:
            continue
        contacts.append({
            'id': str(uuid.uuid4()),
            'customer_id': cust_id,
            'name': name,
            'phone1': row[2],
            'phone2': row[3],
            'position': row[4],
            'gender': row[5],
        })

    print(f'联系人数: {len(contacts)}')

    # ========== 生成SQL ==========
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write('-- ============================================\n')
        f.write('-- 基础客户数据导入（自动生成，勿手动编辑）\n')
        f.write(f'-- 客户: {len(customers)} 条 | 联系人: {len(contacts)} 条\n')
        f.write('-- ============================================\n\n')

        # 清空旧数据
        f.write('DELETE FROM sales_crm_base_contacts;\n')
        f.write('DELETE FROM sales_crm_base_customers;\n\n')

        # 客户 - 分批INSERT（每500条一批）
        write_customer_inserts(f, customers)

        f.write('\n')

        # 联系人 - 分批INSERT
        write_contact_inserts(f, contacts)

    print(f'SQL已生成: {OUTPUT_PATH}')
    print(f'文件大小: {os.path.getsize(OUTPUT_PATH) / 1024 / 1024:.1f} MB')


def write_customer_inserts(f, customers):
    """分批写入客户INSERT语句"""
    BATCH = 500
    for i in range(0, len(customers), BATCH):
        batch = customers[i:i+BATCH]
        f.write(f'-- 客户批次 {i//BATCH + 1}\n')
        f.write('INSERT INTO sales_crm_base_customers '
                '(id, name, legal_representative, social_credit_code, '
                'province, city, county, business_type, customer_type, '
                'detail_address, customer_grade, credit_score, '
                'actual_controller, second_type) VALUES\n')

        lines = []
        for c in batch:
            credit = str(c['credit_score']) if c['credit_score'] is not None else 'NULL'
            line = (f"({escape_sql(c['id'])}, {escape_sql(c['name'])}, "
                    f"{escape_sql(c['legal_representative'])}, {escape_sql(c['social_credit_code'])}, "
                    f"{escape_sql(c['province'])}, {escape_sql(c['city'])}, {escape_sql(c['county'])}, "
                    f"{escape_sql(c['business_type'])}, {escape_sql(c['customer_type'])}, "
                    f"{escape_sql(c['detail_address'])}, {escape_sql(c['customer_grade'])}, "
                    f"{credit}, "
                    f"{escape_sql(c['actual_controller'])}, {escape_sql(c['second_type'])})")
            lines.append(line)

        f.write(',\n'.join(lines))
        f.write(';\n\n')


def write_contact_inserts(f, contacts):
    """分批写入联系人INSERT语句"""
    BATCH = 500
    for i in range(0, len(contacts), BATCH):
        batch = contacts[i:i+BATCH]
        f.write(f'-- 联系人批次 {i//BATCH + 1}\n')
        f.write('INSERT INTO sales_crm_base_contacts '
                '(id, customer_id, name, phone1, phone2, position, gender) VALUES\n')

        lines = []
        for c in batch:
            line = (f"({escape_sql(c['id'])}, {escape_sql(c['customer_id'])}, "
                    f"{escape_sql(c['name'])}, {escape_sql(c['phone1'])}, "
                    f"{escape_sql(c['phone2'])}, {escape_sql(c['position'])}, "
                    f"{escape_sql(c['gender'])})")
            lines.append(line)

        f.write(',\n'.join(lines))
        f.write(';\n\n')


if __name__ == '__main__':
    main()
